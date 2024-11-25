import pandas as pd
import numpy as np
import pickle 
import os
from itertools import combinations
from openpyxl.chart import LineChart, ScatterChart, Reference, trendline
import sys

def determine_workout_splits(df):
    # Check whether the workout splits are already defined
    if os.path.exists("groups.pkl"):
        choice = input("Found existing groups.\n Would you like to use these or continue to define your own? (y/n)")
        if choice == "y":
            with open('groups.pkl', 'rb') as f:
                return pickle.load(f)


    # Otherwise determine the splits by co-occurence and ask the user where co-occurences lead to conflicting groups.
    grouped = df.groupby("title")["exercise_title"].unique()
    cooccurrence = {}
    for exercises in grouped:
        for pair in combinations(exercises, 2):
            pair = tuple(sorted(pair))
            cooccurrence[pair] = cooccurrence.get(pair, 0) + 1

    # Step 3: Convert co-occurrence dict to DataFrame
    cooccurrence_df = pd.DataFrame(
        [(k[0], k[1], v) for k, v in cooccurrence.items()],
        columns=["Exercise 1", "Exercise 2", "Co-occurrences"]
    )

    # Step 4: Apply co-occurrence threshold for grouping
    cooccurrence_threshold = 4  
    filtered_df = cooccurrence_df[cooccurrence_df["Co-occurrences"] >= cooccurrence_threshold]

    # filtered_df = filtered_df.drop(["Co-occurrences"], axis=1)
    print(filtered_df)
    filtered_df = filtered_df.drop_duplicates()
    print(filtered_df)


    groups = {}
    for ex1, ex2, _ in filtered_df.itertuples(index=False):
        ex1_group = ""
        ex2_group = ""
        for group in groups.keys():
            exercises = groups[group]
            if ex1 in exercises:
                ex1_group = group

            if ex2 in exercises:
                ex2_group = group

            if ex1_group != "" and ex2_group != "":
                break


        if ex1_group == ex2_group:
            if ex1_group == "":
                groups["group" + str(len(groups.keys()))] = set([ex1, ex2])
            else:
                # Do nothing, these two exercises are already in the same group
                pass
        elif ex1_group == "" and ex2_group != "":
            # Add ex1 to the same group as ex2
            groups[ex2_group].add(ex1)
        elif ex2_group == "" and ex1_group != "":
            # Add ex2 to the same group as ex1
            groups[ex1_group].add(ex2)
        elif ex2_group != ex1_group:
            # Both exercises are in groups but the groups differ
            print(f"{ex1} and {ex2} appear to be coupled but are already in different groups.")
            set1 = "(" + ', '.join(groups[ex1_group])+")"
            set2 = "(" + ', '.join(groups[ex2_group])+")"
            # Merge groups, or move one exercise to another group
            choice = int(input(f"""
                    1. Move {ex1} to {set2}
                    2. Move {ex2} to {set1}
                    3. Merge the two sets
                    4. Do nothing
                    Choose an option: """))
            if choice == 1:
                groups[ex2_group].add(ex1)
                groups[ex1_group].remove(ex1)
            elif choice == 2:
                groups[ex1_group].add(ex2)
                groups[ex2_group].remove(ex2)
            elif choice == 3:
                merged_name = f"{ex1_group}_{ex2_group}_merged"
                groups[merged_name] = groups[ex1_group].union(groups[ex2_group])
                del groups[ex1_group]
                del groups[ex2_group]
                merged = "(" + ', '.join(groups[ex1_group])+")"
                print(f"Merged: {merged}")
    with open('groups.pkl', 'wb') as f:
        pickle.dump(groups,f)

    return groups

if __name__ == "__main__":
    MASTER_PATH = sys.argv[1]
    NEW_DATA_PATH = sys.argv[2]

    master_df = pd.read_excel(MASTER_PATH, sheet_name="workout_data")
    new_data_df = pd.read_csv(NEW_DATA_PATH)

    # print(pd.concat([master_df, new_data_df],ignore_index=True, axis=0))
    combined_data_df = new_data_df._append([master_df], ignore_index=True)

    prev_num_rows = master_df.shape[0]
    new_num_rows = combined_data_df.shape[0]

    choice = input("Appending this data will result in " + str(new_num_rows) + " additional rows to be inserted. \n Would you like to action this? (y/n)")

    combined_data_df["start_time"] = combined_data_df["start_time"].apply(lambda time: time.split(",")[0])
    combined_data_df["weight_moved"] = combined_data_df["weight_kg"] * combined_data_df["reps"]
    # Reverse order of df to make old data be first
    combined_data_df = combined_data_df[::-1]

    if choice.lower() == "y":
        # combined_data_df.to_csv(MASTER_PATH, index=False)
        # combined_data_df.to_excel(MASTER_PATH, "workout_data", index=False)

        current_data_col = 1
        with pd.ExcelWriter(MASTER_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            combined_data_df.to_excel(writer, sheet_name="workout_data", index=False)
            groups = determine_workout_splits(combined_data_df)
            workbook = writer.book
            data_sheet = workbook.create_sheet("data_sheet")

            for g in range(len(groups.keys())):
                group = list(groups.keys())[g]
                group_sheet = workbook.create_sheet(group)
                chart_count = 0
                exercise_in_group_count = 0
                for exercise in groups[group]:
                    exercise_df = combined_data_df[combined_data_df["exercise_title"] == exercise]
                    print(exercise_df)
                    chart = LineChart()
                    chart.smooth = False
                    chart.title = exercise
                    chart.x_axis.title = "Date"
                    chart.y_axis.title = "Weight Moved"

                    # Add the date and weight data to the data sheet
                    for i, (date, weight) in enumerate(zip(exercise_df["start_time"], exercise_df["weight_moved"]), start=1):
                        data_sheet.cell(row=i, column=current_data_col, value=date)  
                        data_sheet.cell(row=i, column=current_data_col + 1, value=weight)

                    # Create references for x and y axis data
                    xvalues = Reference(data_sheet, min_col=current_data_col, min_row=2, max_row=i)
                    yvalues = Reference(data_sheet, min_col=current_data_col + 1, min_row=2, max_row=i)
                    
                    # Add the data to the chart
                    chart.add_data(yvalues, titles_from_data=False)

                    lobf = trendline.Trendline(dispEq=False, dispRSqr=False)
                    chart.series[0].trendline = lobf
                    chart.series[0].graphicalProperties.line.solidFill = "0000DD"  # Blue line color
                    chart.series[0].graphicalProperties.line.width = 15000 # Line width

                    # Position the chart (e.g., 5 charts per row)
                    col = 2 + (chart_count % 5) * 10
                    row = 2 + (chart_count // 5) * 15

                    chart.legend = None
                    group_sheet.add_chart(chart, f"{group_sheet.cell(row=row, column=col).coordinate}")

                    chart_count += 1
                    current_data_col += 3

                

